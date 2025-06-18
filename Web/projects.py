from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, Response
from . import db
from flask_login import login_required, current_user
from .models import Project, User, SubProject, Note, ChatRoom, Message, Notification
from .functions import has_access_to_project, get_time, get_date, clean_sheet_name, has_redact_rights_to_project
from datetime import datetime, date, timedelta
import json
import csv
from io import BytesIO, TextIOWrapper 
from openpyxl import Workbook
from openpyxl.styles import Font

### ВРЕМЕННОЕ
import sys
###


DB_MAX_INT = 10000000000000000
projects = Blueprint('projects', __name__)

# Показывает все ваши проекты
@projects.route('/projects', methods=["GET", "POST"])
@login_required
def allProjects():
    
    projects = current_user.ownedProjects
    
    # Загружаем из БД уже добавленные пользователем проекты
    added_projects = json.loads(current_user.addedProjects)
    saved_projects = [Project.query.filter_by(id=int(i)).first() for i in added_projects]
    
    if request.method == "POST":
        # Получаем значвения ID проекта из форм страницы
        delete_project_id = request.form.get('delete_project_id')
        add_project_id = request.form.get('add_project')
        
        # Если добавляем проект
        if add_project_id:
            # Проверяем на корректный ввод числа
            try:
                int_add_project_id = int(add_project_id)
            except ValueError:
                # Иначе выводим ошибку и возвращаем страницу без изменений
                flash('Введён некорректный идентификтор')
                return render_template("projects.html", user=current_user, available_other_projects = saved_projects, available_projects = projects, has_access=has_redact_rights_to_project)
            
            # Проверяем на переполнение целого числа в БД
            if int_add_project_id >= DB_MAX_INT:
                flash('Слишком большой номер проекта')
                return render_template("projects.html", user=current_user, available_other_projects = saved_projects, available_projects = projects, has_access=has_redact_rights_to_project)
            
            # Введённые данные корректны, пытаемся запросить нужный проект
            project = Project.query.filter_by(id=int(int_add_project_id)).first()
            if not project:
                flash('Проекта с таким ID не существует')
                
            # Проект существует, проверяем есть ли доступ
            elif has_access_to_project(user=current_user, project=project):
                if  project.owner_id == current_user.id:
                    flash('Вы владелец этого проекта')
                elif add_project_id in added_projects:
                    flash('Вы уже сохранили данный проект')
                else:
                    # Долбавляем проект в сохранённые у пользователя
                    added_projects.append(add_project_id)
                    saved_projects.append(project)
            else:
                flash('Доступ к запрашиваемому проекту запрещён')
                
        # Удаление проекта из добавленных           
        elif delete_project_id:
            if delete_project_id in added_projects:
                # Проверяем на корректный ввод числа
                try:
                    int_delete_project_id = int(delete_project_id)
                except ValueError:
                    flash('Введён некорректный идентификтор')
                    return render_template("projects.html", user=current_user, available_other_projects = saved_projects, available_projects = projects, has_access=has_redact_rights_to_project)
                if int_delete_project_id >= DB_MAX_INT:
                    flash('Слишком большой номер проекта')
                    return render_template("projects.html", user=current_user, available_other_projects = saved_projects, available_projects = projects, has_access=has_redact_rights_to_project)
                
                # Введённые данные корректны, удаляем введённый проект
                project = Project.query.filter_by(id=int_delete_project_id).first()
                added_projects.remove(delete_project_id)
                saved_projects.remove(project)
        
        # Сохраняем проект в поля пользователя  
        current_user.addedProjects = json.dumps(added_projects)
        # Сохраняем изменения в БД
        db.session.commit()
        return render_template("projects.html", user=current_user, available_other_projects = saved_projects, available_projects = projects, has_access=has_redact_rights_to_project)
    
    return render_template("projects.html", user=current_user, available_projects = projects, has_access=has_redact_rights_to_project, available_other_projects = saved_projects)

# Показывает все проекты
@projects.route('/admin_projects', methods=["GET"])
@login_required
def adminProjects():
    
    if current_user.is_admin != "True":
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return render_template("forbidden.html", user=current_user)
    
    projects = Project.query.all()
    return render_template("projects_admin.html", user=current_user, available_projects = projects, has_access=has_redact_rights_to_project)

# Показывает все ваши проекты
@projects.route('/admin_users', methods=["GET", "POST"])
@login_required
def adminUsers():
    
    if current_user.is_admin != "True":
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return render_template("forbidden.html", user=current_user)
    
    users = User.query.all()
    
    if request.method == "GET":
        return render_template("admin_users.html", user=current_user, users = users, get_date=get_date)
    
    user_id = request.form.get('user_id')
    user = User.query.filter_by(id=int(user_id)).first()
    
    if not user:
        return render_template("admin_users.html", user=current_user, users = users, get_date=get_date)
    
    if user.id == current_user.id:
        flash("Невозможно совершать действия над своим аккаунтом", category="error")
        return render_template("admin_users.html", user=current_user, users = users, get_date=get_date)
    elif user.id == 1:
        flash("Невозможно совершить действие над аккаунтом главного администратора", category="error")
        return render_template("admin_users.html", user=current_user, users = users, get_date=get_date)
    
    if 'role_button' in request.form:
        user.is_admin = ["True", "False"][user.is_admin == "True"]
        
    elif 'delete_button' in request.form:
        db.session.delete(user)
        
    # В БУДУЩЕМ УДАЛЯТЬ ПРОЕКТЫ ПОЛЬЗОВАТЕЛЯ №№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№№
    db.session.commit()
    users = User.query.all()
    return render_template("admin_users.html", user=current_user, users = users, get_date=get_date)


# Управляет функционалом по редакитированию информации о конкретном проекте
@projects.route('/projects/<int:index>/redact', methods=["POST", "GET"])
@login_required
def redactProject(index):
    project = Project.query.filter_by(id=index).first()
    
    # В случае некорректной информации останавливаем выполнение
    if not project:
        return redirect(url_for('projects.allProjects'))

    # Если у пользователя нет доступа - останавливаем дальнейшую работу
    if not has_access_to_project(user=current_user, project=project):
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return render_template("forbidden.html", user=current_user)
    
    if not has_redact_rights_to_project(user=current_user, project=project):
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return redirect(url_for('projects.showProject', index=index))
    
    # Получаем актуальную информацию из БД
    allowed_users_list = json.loads(project.allowedUsers)
    
    if request.method == "GET":
        allowed_users = [[User.query.filter_by(id=int(i[0])).first(), i[1]] for i in allowed_users_list]
        return render_template("redact_project.html", project=project, user=current_user, allowed_users=allowed_users)
    
    # Получение информации из запроса
    name = request.form.get('name')
    shortDescription = request.form.get('shortDescription')
    fullDescription = request.form.get('fullDescription')
    goal = request.form.get('goal')
    delete_word = request.form.get('delete')
    user_id = request.form.get('user_id')
    to_delete_user_id = request.form.get('to_delete_user_id')
 
    # Удаление проекта
    if delete_word == "УДАЛИТЬ":
        project.delete(db)
        db.session.commit()
        return redirect(url_for('projects.allProjects'))
    
    if 'save_button' in request.form:
        project.name = name
        project.fullDescription = fullDescription
        project.shortDescription = shortDescription
        project.goal = goal
        
    elif to_delete_user_id:
        to_pop_index = -1
        for num, record in enumerate(allowed_users_list):
            if record[0] == to_delete_user_id:
                to_pop_index = num
                break

        if to_pop_index == -1:
            return redirect(url_for('projects.allProjects'))
        
        allowed_users_list.pop(to_pop_index)
        project.allowedUsers = json.dumps(allowed_users_list)
        
    # Изменение состояния проекта
    elif 'done' in request.form:
        if project.done == "True":
            project.done = "False"
        else:
            project.done = "True"
            
            
    # Добавление пользователей к проекту
    elif user_id:
        user_to_add = User.query.filter_by(id=user_id).first()
        if not user_to_add:
            flash("Пользователь с таким ID не найден", category="error")
    
        elif project.owner_id == user_to_add.id:
            flash("Вы не можете добавить владельца", category="error")
            
        else:
            # Проверяем, если ли уже у запрашиваемого пользователя доступ к проекту
            for record in allowed_users_list:
                if user_id == record[0]:
                    flash("Пользователю с таким ID уже предоставлен доступ", category="error")
                    break
            else:
                allowed_users_list.append([user_id, 0])
                project.allowedUsers = json.dumps(allowed_users_list)   
                
    elif 'toggle_access'in request.form:            
        to_change_access_user_id = request.form.get('toggle_access')

        for record in allowed_users_list:
            if record[0] == to_change_access_user_id:
                record[1] = not(record[1])
                break
        
        project.allowedUsers = json.dumps(allowed_users_list)
        
    db.session.commit()
    
    allowed_users = [[User.query.filter_by(id=int(i[0])).first(), i[1]] for i in allowed_users_list]
    return render_template("redact_project.html", project=project, user=current_user, allowed_users=allowed_users)

# Описывает создание проекта
@projects.route('/new_project', methods=["GET", "POST"])
@login_required
def newProject():
    if request.method == "GET":
        return render_template("new_project.html", user=current_user)
    
    elif request.method == "POST":
        name = request.form.get('name')
        shortDescription = request.form.get('shortDescription')

        new_project = Project(name=name, shortDescription=shortDescription, owner_id=current_user.id)
        db.session.add(new_project)
        db.session.commit()
        
        new_chat_room = ChatRoom(parent_project_id=new_project.id)
        db.session.add(new_chat_room)
        db.session.commit()

        return redirect(url_for('projects.allProjects'))

# Описывает логику просмотра проекта
@projects.route('/projects/<int:index>', methods=["POST", "GET"])
@login_required
def showProject(index):
    project = Project.query.filter_by(id=index).first()
    if not project:
        return redirect(url_for('projects.allProjects'))
    
    if not has_access_to_project(user=current_user, project=project):
        return redirect(url_for('projects.allProjects'))
    
    subprojects = project.subprojects
    
    if not has_redact_rights_to_project(user=current_user, project=project):
        return render_template("show_project.html", project=project, user=current_user, subprojects=subprojects)
    
    if request.method == "GET":
        return render_template("show_project.html", project=project, user=current_user, subprojects=subprojects)
    
    to_create_subproject_name = request.form.get('to_create_subproject_name')
    
    if to_create_subproject_name:
        subproject = SubProject(name = to_create_subproject_name, parent_id = project.id)
        db.session.add(subproject)
        subprojects.append(subproject)
        
    db.session.commit()
    return render_template("show_project.html", project=project, user=current_user, subprojects=subprojects)

# Описывает редактирования раздела
@projects.route('/projects/<int:index>/<int:subproject>/redact', methods=["POST", "GET"])
@login_required
def redactSubProject(index, subproject):
    project = Project.query.filter_by(id=index).first()
    subproject = SubProject.query.filter_by(id=subproject).first()    
    # В случае некорректной информации останавливаем выполнение
    if not project or not subproject:
        return redirect(url_for('projects.allProjects'))
    
    if not has_access_to_project(user=current_user, project=project): 
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return render_template("forbidden.html", user=current_user)
    
    if not has_redact_rights_to_project(user=current_user, project=project):
        flash("У вас отсутствуют права для просмотра данной страницы", category="error")
        return redirect(url_for('projects.showProject', index=index))

   
    if request.method == "POST":
        name = request.form.get('name')
        description = request.form.get('description')
        delete_word = request.form.get('delete')
        is_subproject_done_id = request.form.get('done')

        if delete_word == "УДАЛИТЬ":
            subproject.delete(db)
            project.update_progress(db)   
            db.session.commit()
            
            return redirect(url_for('projects.showProject', index=project.id, **request.args))
        
        if name:
            subproject.name = name
            subproject.shortDescription = description

        if is_subproject_done_id:
            if subproject.done == "True":
                subproject.done = "False"
            else:
                subproject.done = "True"
            
        db.session.commit()    
        
    return render_template("redact_subproject.html", project=project, user=current_user, subproject=subproject)
    
@projects.route('/projects/<int:index>/<int:subproject>', methods=["GET", "POST"])
@login_required
def showSubProject(index, subproject):
    if request.method == "POST":
        
        project = Project.query.filter_by(id=index).first()
        subproject = SubProject.query.filter_by(id=subproject).first()
        
        if not has_access_to_project(user=current_user, project=project):
            return render_template("forbidden.html", user=current_user)
        
        if not has_redact_rights_to_project(user=current_user, project=project):
            return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
    
        create_note = request.form.get('create_note')
        title = request.form.get('title')
        content = request.form.get('content')
        note_id = request.form.get('note_id')
        
        # DEBUGGING TOOL  print(request.form, file=sys.stdout) ==============================================================
        
        if not note_id and not create_note:
            return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
        
        if 'create_note' in request.form:
            note = Note(parent_id=subproject.id, title=create_note, status="ready")
            db.session.add(note)
            project.update_progress(db)   
            db.session.commit()
            return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
        
        note = Note.query.filter_by(id=int(note_id)).first()
        
        if not note:
            return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
        
        if 'save_button' in request.form:
            note.title = title
            note.content = content
        
        elif 'status_button' in request.form:
            note.next_status()
              
        elif 'comments_button'in request.form:

            if not note.chatRoom:
                chat_room = ChatRoom(parent_note_id=note.id)
                db.session.add(chat_room)
                db.session.commit()
            else:
                chat_room = note.chatRoom
                    
            return redirect(url_for('projects.accessNoteChat', project_id=project.id, subproject_id=subproject.id, note_id=note.id, **request.args))  
              
        elif 'delete_button' in request.form:
            
            note.delete(db)
            project.update_progress(db)    
            
        elif 'planning_button' in request.form:
            
            return redirect(url_for('projects.manageNote', project_id=project.id, subproject_id=subproject.id, note_id=note.id, **request.args))  
                
        db.session.commit()
        return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
    
    elif request.method == "GET":
        
        project = Project.query.filter_by(id=index).first()
        subproject = SubProject.query.filter_by(id=subproject).first()
        
        if not has_access_to_project(user=current_user, project=project):
            return render_template("forbidden.html", user=current_user)
        
        return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
    
    return render_template("forbidden.html", user=current_user)

# Обрабатывает управление конкретной задачей
@projects.route('/projects/<int:project_id>/<int:subproject_id>/<int:note_id>/redact', methods=["GET", "POST"])
@login_required
def manageNote(project_id, subproject_id, note_id):
    
    project = Project.query.filter_by(id=project_id).first()
    subproject = SubProject.query.filter_by(id=subproject_id).first()
    
    if not has_access_to_project(user=current_user, project=project):
        return render_template("forbidden.html", user=current_user)    

    if not has_redact_rights_to_project(user=current_user, project=project):   
        return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)
    
    note = Note.query.filter_by(id=int(note_id)).first()
    
    if not note:
        return render_template("show_subproject.html", project=project, user=current_user, subproject=subproject, get_date=get_date)   
    
    notification = note.notification
            
    if request.method == "GET":
        return render_template("redact_note.html", project=project, user=current_user, subproject=subproject, note=note, notification=notification, get_date=get_date)
    
    elif request.method == "POST":
            
        progress = request.form.get('progress')
        progress_coefficient = request.form.get('progress_coefficient')
        planned_at = request.form.get('planned_at')
        # ADD VALIDATION 1 - 100
        if 'save_button' in request.form:
            note.progress = progress
            note.progress_coefficient = progress_coefficient
            note.planned_at = planned_at
            
            project.update_progress(db)
            
        elif 'change_notification_status' in request.form:   
            note.notifications_enabled = ["False", "True"][note.notifications_enabled == "False"]   
            
        elif 'create_notification' in request.form:  
            if note.planned_at:
                notif_offset = request.form.get('notification_time') 
                time_delta = timedelta(hours=int(notif_offset))
                notification_date = (datetime.strptime(note.planned_at, "%Y-%m-%d") - time_delta).date()
                if notification:
                    notification.planned_at = str(notification_date)
                else:
                    notification = Notification(planned_at=notification_date, parent_id=note.id)
                    db.session.add(notification)
                    
        db.session.commit()
        return render_template("redact_note.html", project=project, user=current_user, subproject=subproject, note=note, notification = notification, get_date=get_date)
    
    return render_template("forbidden.html", user=current_user)

# Обрабатывает запросы по адресу /other_projects с методами Get и Post 
@projects.route('/other_projects', methods=["GET", "POST"])
@login_required
def showOtherProjects():
    # Управялет логикой сохранения проектов других пользователей
    # Только аутентифицированный пользователь имеет доступ к функционалу данной подпрограммы
    # Метод POST подаётся при нажатии на кнопку
    if request.method == "POST":
        
        # Получаем значвения ID проекта из форм страницы
        delete_project_id = request.form.get('delete_project_id')
        add_project_id = request.form.get('add_project')
        
        # Загружаем из БД уже добавленные пользователем проекты
        added_projects = json.loads(current_user.addedProjects)
        saved_projects = [Project.query.filter_by(id=int(i)).first() for i in added_projects]
        
        # Если добавляем проект
        if add_project_id:
            # Проверяем на корректный ввод числа
            try:
                int_add_project_id = int(add_project_id)
            except ValueError:
                # Иначе выводим ошибку и возвращаем страницу без изменений
                flash('Введён некорректный идентификтор')
                return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
            # Проверяем на переполнение целого числа в БД
            if int_add_project_id >= DB_MAX_INT:
                flash('Слишком большой номер проекта')
                return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
            
            # Введённые данные корректны, пытаемся запросить нужный проект
            project = Project.query.filter_by(id=int(int_add_project_id)).first()
            if not project:
                flash('Проекта с таким ID не существует')
            # Проект существует, проверяем есть ли доступ
            elif has_access_to_project(user=current_user, project=project):
                if  project.owner_id == current_user.id:
                    flash('Вы владелец этого проекта')
                elif add_project_id in added_projects:
                    flash('Вы уже сохранили данный проект')
                else:
                    # Долбавляем проект в сохранённые у пользователя
                    added_projects.append(add_project_id)
                    saved_projects.append(project)
            else:
                flash('Доступ к запрашиваемому проекту запрещён')
                
        # Удаление проекта из добавленных           
        elif delete_project_id:
            if delete_project_id in added_projects:
                # Проверяем на корректный ввод числа
                try:
                    int_delete_project_id = int(delete_project_id)
                except ValueError:
                    flash('Введён некорректный идентификтор')
                    return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
                if int_delete_project_id >= DB_MAX_INT:
                    flash('Слишком большой номер проекта')
                    return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
                
                # Введённые данные корректны, удаляем введённый проект
                project = Project.query.filter_by(id=int_delete_project_id).first()
                added_projects.remove(delete_project_id)
                saved_projects.remove(project)
        
        # Сохраняем проект в поля пользователя  
        current_user.addedProjects = json.dumps(added_projects)
        # Сохраняем изменения в БД
        db.session.commit()
        return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
    
    # При переходе по адресу
    elif request.method == "GET":   
        # При переходе по адресу
        # Загружаем из БД уже добавленные пользователем проекты
        added_projects = json.loads(current_user.addedProjects)
        saved_projects = [Project.query.filter_by(id=int(i)).first() for i in added_projects]
        
        # Отображаем страницу
        return render_template("other_projects.html", user=current_user, available_projects = saved_projects)
    
    
# Управляет проектным чатом
@projects.route('/projects/<int:index>/chat', methods=["GET", "POST"])
@login_required
def accessProjectChat(index):
    project = Project.query.filter_by(id=index).first()
    chat_room = project.chatRoom

    # Если у пользователя нет доступа - останавливаем дальнейшую работу
    if not has_access_to_project(user=current_user, project=project):
        return render_template("forbidden.html", user=current_user)
    
    if request.method == "GET":
        return render_template("show_project_chat.html", project=project, user=current_user, chat_room=chat_room, get_time=get_time)
    
    # POST
    create_message = request.form.get('create_message')
    
    if create_message:
        message = Message(parent_id=chat_room.id, text=create_message, created_at=datetime.now().strftime('%Y-%m-%d-%H-%M'), \
                            author_name = current_user.firstName, author_id = current_user.id)
        db.session.add(message)
        db.session.commit()
        
    return render_template("show_project_chat.html", project=project, user=current_user, chat_room=chat_room, get_time=get_time)

# Управляет комментариями к задачам
@projects.route('/projects/<int:project_id>/<int:subproject_id>/<int:note_id>/comments', methods=["GET", "POST"])
@login_required
def accessNoteChat(project_id, note_id, subproject_id):
    project = Project.query.filter_by(id=project_id).first()
    subproject = SubProject.query.filter_by(id=subproject_id).first()
    note = Note.query.filter_by(id=note_id).first()
    chat_room = note.chatRoom

    # Если у пользователя нет доступа - останавливаем дальнейшую работу
    if not has_access_to_project(user=current_user, project=project):
        return render_template("forbidden.html", user=current_user)
    
    if request.method == "GET":
        return render_template("show_note_chat.html", subproject=subproject, user=current_user, chat_room=chat_room, note=note, project=project, \
            get_time=get_time) 
    # POST
    create_message = request.form.get('create_message')
    
    if create_message:
        message = Message(parent_id=chat_room.id, text=create_message, created_at=datetime.now().strftime('%Y-%m-%d-%H-%M'), \
                            author_name = current_user.firstName, author_id = current_user.id)
        db.session.add(message)
        db.session.commit()
        
    return render_template("show_note_chat.html", subproject=subproject, user=current_user, chat_room=chat_room, note=note, project=project, \
            get_time=get_time) 
    
# Управляет статистикой и экспортом
@projects.route('/projects/<int:index>/stats', methods=["POST", "GET"])
@login_required
def showStats(index):
    project = Project.query.filter_by(id=index).first()
    
    # В случае некорректной информации останавливаем выполнение
    if not project:
        return redirect(url_for('projects.allProjects'))

    # Если у пользователя нет доступа - останавливаем дальнейшую работу
    if not has_access_to_project(user=current_user, project=project):
        return render_template("forbidden.html", user=current_user)
    
    subprojects = project.subprojects
    subprojects_num = len(subprojects)
    
    tasks_num = 0
    tasks_ready = 0
    tasks_done = 0
    tasks_abandoned = 0
    tasks_in_progress = 0
    
    due_3d = []
    due_week = []
    due_month = []
    
    for subproject in subprojects:
        for note in subproject.notes:
            if note.status == "in_progress":
                tasks_in_progress += 1
            elif note.status == "ready":
                tasks_ready += 1
            elif note.status == "done":
                tasks_done += 1
            elif note.status == "abandoned":
                tasks_abandoned += 1
            if note.planned_at and note.status != 'done':
                planned_time = date(*[int(i) for i in note.planned_at.split("-")])
                current_time = date.today()     
                timedelta = planned_time - current_time
                if timedelta.days < 0:
                    continue
                elif timedelta.days <= 3:
                    due_3d.append(note)
                elif timedelta.days <= 7:
                    due_week.append(note)
                elif timedelta.days <= 30: 
                    due_month.append(note)
                         
    tasks_num = tasks_ready + tasks_done + tasks_abandoned + tasks_in_progress
    
    if request.method == "GET":
        return render_template("show_stats.html", project=project, user=current_user, due_3d = due_3d, due_week=due_week, due_month=due_month, \
            subprojects_num=subprojects_num, tasks_in_progress=tasks_in_progress, tasks_abandoned=tasks_abandoned, tasks_done=tasks_done, tasks_ready=tasks_ready, tasks_num=tasks_num, get_date=get_date)
    
    return render_template("show_stats.html", project=project, user=current_user)

# Экспортирует проект
@projects.route('/projects/<int:project_id>/export_csv', methods=["POST"])
@login_required
def exportProjectCSV(project_id):
    # Получаем проект и его подпроекты
    project = Project.query.filter_by(id=project_id).first()
    # В случае некорректной информации останавливаем выполнение
    if not project:
        return redirect(url_for('projects.allProjects'))
    
    subprojects = project.subprojects

    # Создаем CSV в памяти
    csv_buffer = BytesIO()
    csv_buffer.write(b'\xef\xbb\xbf')
    
    # Оборачиваем BytesIO в TextIOWrapper с указанием кодировки
    text_buffer = TextIOWrapper(
        csv_buffer,
        encoding='utf-8',
        write_through=True  # Автоматическая запись данных в BytesIO
    )
    
    writer = csv.writer(text_buffer, lineterminator = '\n')
        
    # Заголовки CSV
    writer.writerow([
        'Project ID', 
        'Project Name', 
        'Full Description', 
        'Parent Project ID'
    ])
    
    # Данные основного проекта
    writer.writerow([
        project.id,
        project.name,
        project.fullDescription,
        ''  # У основного проекта нет родителя
    ])
    
    # Данные подпроектов
    for sub in subprojects:
        writer.writerow([
            sub.id,
            sub.name,
            sub.shortDescription,
            sub.parent_id
        ])
    
    # Критически важные действия:
    text_buffer.flush()  # Форсируем запись всех данных
    csv_buffer = text_buffer.detach()  # Отсоединяем BytesIO от TextIOWrapper
    csv_buffer.seek(0)  # Перемещаем указатель в начало
       
    return Response(
        csv_buffer,
        mimetype="text/csv",
        headers={"Content-disposition":
                 "attachment; filename=project_data.csv"})


# Экспортирует проект
@projects.route('/projects/<int:project_id>/export_excel', methods=["POST"])
@login_required
def exportProjectExcel(project_id):
    # Получаем проект и его подпроекты
    project = Project.query.filter_by(id=project_id).first()
    # В случае некорректной информации останавливаем выполнение
    if not project:
        return redirect(url_for('projects.allProjects'))
    
    subprojects = project.subprojects
    
    # Создаем книгу Excel
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    # Лист основного проекта
    main_ws = wb.create_sheet(title=clean_sheet_name(project.name))
    main_ws.append(['ID', 'Project Name', 'Description'])
    main_ws.append([project.id, project.name, project.fullDescription])
    
    # Форматирование заголовков
    bold_font = Font(bold=True)
    for row in main_ws.iter_rows(max_row=1):
        for cell in row:
            cell.font = bold_font

    # Листы для подпроектов
    for sub in subprojects:
        sheet_name = clean_sheet_name(f"Sub: {sub.name}")
        sub_ws = wb.create_sheet(title=sheet_name)
        
        # Заголовки подпроекта
        sub_ws.append(['Subproject ID', 'Name', 'Description'])
        sub_ws.append([sub.id, sub.name, sub.shortDescription])
        
        # Задачи подпроекта
        sub_ws.append([])  # Пустая строка
        sub_ws.append(['Tasks'])
        sub_ws.append(['Task ID', 'Name', 'Description', 'Status'])
        
        tasks = Note.query.filter_by(parent_id=sub.id).all()
        for task in tasks:
            sub_ws.append([task.id, task.title, task.content, task.status])
        
        # Форматирование и автоподбор ширины
        for column in sub_ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sub_ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Жирный шрифт для заголовков
        for row in sub_ws.iter_rows(max_row=3):
            for cell in row:
                cell.font = bold_font

    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'project_{project_id}_export.xlsx'
    )